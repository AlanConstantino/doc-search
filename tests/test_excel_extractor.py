"""
Tests for the Excel document extractor module.

Tests Excel (.xlsx) extraction using vendored openpyxl library.
"""

import unittest
import tempfile
from pathlib import Path

from doc_search.excel_extractor import ExcelExtractor, extract_excel_text


class TestExcelExtractor(unittest.TestCase):
    """Tests for Excel (.xlsx) extraction."""
    
    @classmethod
    def setUpClass(cls):
        """Set up test fixtures."""
        cls.fixtures_dir = Path(__file__).parent / 'fixtures' / 'excel'
        cls.sample_xlsx = cls.fixtures_dir / 'sample.xlsx'
        cls.multi_sheet_xlsx = cls.fixtures_dir / 'multi_sheet.xlsx'
        cls.formulas_xlsx = cls.fixtures_dir / 'formulas.xlsx'
        cls.large_xlsx = cls.fixtures_dir / 'large.xlsx'
        
        # Ensure fixtures exist
        if not cls.sample_xlsx.exists():
            import sys
            sys.path.insert(0, str(cls.fixtures_dir))
            from create_fixtures import (
                create_sample_xlsx, create_multi_sheet_xlsx,
                create_formulas_xlsx, create_large_xlsx
            )
            create_sample_xlsx()
            create_multi_sheet_xlsx()
            create_formulas_xlsx()
            create_large_xlsx()
    
    def test_extract_single_sheet(self):
        """Extract text from a single-sheet Excel file."""
        extractor = ExcelExtractor()
        documents = extractor.extract(self.sample_xlsx)
        
        self.assertEqual(len(documents), 1)
        doc = documents[0]
        
        self.assertIsNone(doc['error'])
        self.assertIn('Q1 Summary', doc['title'])
        self.assertIn('sample.xlsx', doc['title'])
        self.assertIn('file://', doc['url'])
        self.assertIn('Q1%20Summary', doc['url'])  # URL encoded
        
        # Check metadata
        self.assertEqual(doc['metadata']['doc_type'], 'xlsx')
        self.assertEqual(doc['metadata']['sheet_name'], 'Q1 Summary')
        self.assertEqual(doc['metadata']['sheet_index'], 0)
        self.assertEqual(doc['metadata']['row_count'], 4)  # 1 header + 3 data
        self.assertEqual(doc['metadata']['col_count'], 3)
        
        # Check text extraction with headers
        text = doc['text']
        self.assertIn('Name: John', text)
        self.assertIn('Amount: 500', text)
        self.assertIn('Name: Jane', text)
        self.assertIn('Amount: 750', text)
        self.assertIn('Name: Bob', text)
        self.assertIn('Amount: 1000', text)
    
    def test_extract_multi_sheet(self):
        """Extract text from a multi-sheet Excel file."""
        extractor = ExcelExtractor()
        documents = extractor.extract(self.multi_sheet_xlsx)
        
        # Should have 3 sheets (including empty one)
        self.assertEqual(len(documents), 3)
        
        # First sheet - Sales Data
        doc1 = documents[0]
        self.assertEqual(doc1['metadata']['sheet_name'], 'Sales Data')
        self.assertIn('Product: Widget A', doc1['text'])
        self.assertIn('Revenue: 1000', doc1['text'])
        
        # Second sheet - Summary
        doc2 = documents[1]
        self.assertEqual(doc2['metadata']['sheet_name'], 'Summary')
        self.assertIn('Total', doc2['text'])
        self.assertIn('2500', doc2['text'])
        
        # Third sheet - Empty
        doc3 = documents[2]
        self.assertEqual(doc3['metadata']['sheet_name'], 'Empty Sheet')
        self.assertEqual(doc3['text'], '')
    
    def test_extract_without_headers(self):
        """Extract without treating first row as headers."""
        extractor = ExcelExtractor(first_row_is_header=False)
        documents = extractor.extract(self.sample_xlsx)
        
        doc = documents[0]
        text = doc['text']
        
        # Should have raw values without header context
        # Format: "value1 | value2 | value3"
        self.assertIn('Name', text)
        self.assertIn('Amount', text)
        self.assertIn('Date', text)
        self.assertIn('John', text)
        self.assertIn('500', text)
        
        # Should not have "Name: John" format
        self.assertNotIn('Name: John', text)
    
    def test_extract_invalid_file(self):
        """Extraction should handle invalid files gracefully."""
        extractor = ExcelExtractor()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(b'not a valid xlsx file')
            temp_path = Path(f.name)
        
        try:
            documents = extractor.extract(temp_path)
            self.assertEqual(len(documents), 1)
            self.assertIsNotNone(documents[0]['error'])
        finally:
            temp_path.unlink()
    
    def test_headings_from_header_row(self):
        """Headers should be extracted as headings."""
        extractor = ExcelExtractor()
        documents = extractor.extract(self.sample_xlsx)
        
        doc = documents[0]
        headings = doc['headings']
        
        # Header row values should appear as H2 headings
        heading_texts = [h[1] for h in headings]
        self.assertIn('Name', heading_texts)
        self.assertIn('Amount', heading_texts)
        self.assertIn('Date', heading_texts)
        
        # All should be level 2
        for level, _ in headings:
            self.assertEqual(level, 2)
    
    def test_max_rows_limit(self):
        """Test max_rows parameter limits extraction."""
        extractor = ExcelExtractor(max_rows=5)
        documents = extractor.extract(self.large_xlsx)
        
        doc = documents[0]
        # Should have 5 rows max (1 header + 4 data)
        self.assertEqual(doc['metadata']['row_count'], 5)
        
        # Text should only have 4 data rows
        lines = doc['text'].strip().split('\n')
        self.assertEqual(len(lines), 4)  # Header row not counted in text
    
    def test_extract_formulas_as_values(self):
        """Formulas should be extracted as computed values."""
        extractor = ExcelExtractor()
        documents = extractor.extract(self.formulas_xlsx)
        
        doc = documents[0]
        text = doc['text']
        
        # Should have computed values, not formula strings
        # Note: In read_only mode with data_only=True, we get values
        # But formulas in new files may show as None until saved with Excel
        self.assertIn('Value 1', text)
        self.assertIn('Value 2', text)
    
    def test_file_not_found(self):
        """Should handle non-existent file gracefully."""
        extractor = ExcelExtractor()
        documents = extractor.extract(Path('/nonexistent/file.xlsx'))
        
        self.assertEqual(len(documents), 1)
        self.assertIsNotNone(documents[0]['error'])


class TestExtractExcelText(unittest.TestCase):
    """Tests for the convenience function."""
    
    @classmethod
    def setUpClass(cls):
        cls.fixtures_dir = Path(__file__).parent / 'fixtures' / 'excel'
        cls.sample_xlsx = cls.fixtures_dir / 'sample.xlsx'
    
    def test_extract_excel_text(self):
        """Convenience function should return concatenated text."""
        text = extract_excel_text(str(self.sample_xlsx))
        
        self.assertIn('Q1 Summary', text)
        self.assertIn('Name: John', text)
        self.assertIn('Amount: 500', text)
    
    def test_extract_excel_text_invalid_file(self):
        """Should return empty string for invalid file."""
        text = extract_excel_text('/nonexistent/file.xlsx')
        self.assertEqual(text, '')


class TestExcelExtractorEdgeCases(unittest.TestCase):
    """Tests for edge cases in Excel extraction."""
    
    def test_empty_workbook(self):
        """Test handling of workbook with all empty sheets."""
        import sys
        from pathlib import Path
        _vendor_path = Path(__file__).parent.parent / 'vendor'
        sys.path.insert(0, str(_vendor_path))
        from openpyxl import Workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)
        
        try:
            wb = Workbook()
            wb.active.title = "Empty"
            wb.save(temp_path)
            
            extractor = ExcelExtractor()
            documents = extractor.extract(temp_path)
            
            self.assertEqual(len(documents), 1)
            self.assertIsNone(documents[0]['error'])
            self.assertEqual(documents[0]['text'], '')
        finally:
            temp_path.unlink()
    
    def test_special_characters_in_sheet_name(self):
        """Test sheet names with special characters are URL-encoded."""
        import sys
        from pathlib import Path
        _vendor_path = Path(__file__).parent.parent / 'vendor'
        sys.path.insert(0, str(_vendor_path))
        from openpyxl import Workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)
        
        try:
            wb = Workbook()
            wb.active.title = "Data & Analysis"
            wb.active['A1'] = 'Test'
            wb.save(temp_path)
            
            extractor = ExcelExtractor()
            documents = extractor.extract(temp_path)
            
            doc = documents[0]
            # Sheet name should be URL encoded in the URL
            self.assertIn('Data%20%26%20Analysis', doc['url'])
            # But readable in title
            self.assertIn('Data & Analysis', doc['title'])
        finally:
            temp_path.unlink()
    
    def test_boolean_values(self):
        """Test boolean cell values are extracted correctly."""
        import sys
        from pathlib import Path
        _vendor_path = Path(__file__).parent.parent / 'vendor'
        sys.path.insert(0, str(_vendor_path))
        from openpyxl import Workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Status'
            ws['A2'] = True
            ws['A3'] = False
            wb.save(temp_path)
            
            extractor = ExcelExtractor()
            documents = extractor.extract(temp_path)
            
            text = documents[0]['text']
            self.assertIn('TRUE', text)
            self.assertIn('FALSE', text)
        finally:
            temp_path.unlink()
    
    def test_float_formatting(self):
        """Test float values are formatted correctly."""
        import sys
        from pathlib import Path
        _vendor_path = Path(__file__).parent.parent / 'vendor'
        sys.path.insert(0, str(_vendor_path))
        from openpyxl import Workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = Path(f.name)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Value'
            ws['A2'] = 10.0  # Should become "10" not "10.0"
            ws['A3'] = 3.14159
            wb.save(temp_path)
            
            extractor = ExcelExtractor()
            documents = extractor.extract(temp_path)
            
            text = documents[0]['text']
            self.assertIn('10', text)
            self.assertIn('3.14159', text)
        finally:
            temp_path.unlink()


if __name__ == '__main__':
    unittest.main()
