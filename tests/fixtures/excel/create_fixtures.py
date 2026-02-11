#!/usr/bin/env python3
"""
Create test fixtures for Excel extractor tests using openpyxl.
"""

import sys
from pathlib import Path

# Add vendor directory to path
_vendor_path = Path(__file__).parent.parent.parent.parent / 'vendor'
sys.path.insert(0, str(_vendor_path))

from openpyxl import Workbook

FIXTURE_DIR = Path(__file__).parent


def create_sample_xlsx():
    """Create a sample Excel file with one sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 Summary"
    
    # Header row
    ws['A1'] = 'Name'
    ws['B1'] = 'Amount'
    ws['C1'] = 'Date'
    
    # Data rows
    ws['A2'] = 'John'
    ws['B2'] = 500
    ws['C2'] = '2024-01-15'
    
    ws['A3'] = 'Jane'
    ws['B3'] = 750
    ws['C3'] = '2024-01-16'
    
    ws['A4'] = 'Bob'
    ws['B4'] = 1000
    ws['C4'] = '2024-01-17'
    
    wb.save(FIXTURE_DIR / 'sample.xlsx')
    print(f"Created: {FIXTURE_DIR / 'sample.xlsx'}")


def create_multi_sheet_xlsx():
    """Create an Excel file with multiple sheets."""
    wb = Workbook()
    
    # First sheet - Sales Data
    ws1 = wb.active
    ws1.title = "Sales Data"
    ws1['A1'] = 'Product'
    ws1['B1'] = 'Revenue'
    ws1['A2'] = 'Widget A'
    ws1['B2'] = 1000
    ws1['A3'] = 'Widget B'
    ws1['B3'] = 1500
    
    # Second sheet - Summary
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = 'Total'
    ws2['B1'] = 2500
    ws2['A2'] = 'Average'
    ws2['B2'] = 1250
    
    # Third sheet - Empty
    ws3 = wb.create_sheet("Empty Sheet")
    
    wb.save(FIXTURE_DIR / 'multi_sheet.xlsx')
    print(f"Created: {FIXTURE_DIR / 'multi_sheet.xlsx'}")


def create_formulas_xlsx():
    """Create an Excel file with formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"
    
    ws['A1'] = 'Value 1'
    ws['B1'] = 'Value 2'
    ws['C1'] = 'Sum'
    
    ws['A2'] = 10
    ws['B2'] = 20
    ws['C2'] = '=A2+B2'  # Formula
    
    ws['A3'] = 30
    ws['B3'] = 40
    ws['C3'] = '=A3+B3'  # Formula
    
    wb.save(FIXTURE_DIR / 'formulas.xlsx')
    print(f"Created: {FIXTURE_DIR / 'formulas.xlsx'}")


def create_large_xlsx():
    """Create a larger Excel file for testing max_rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Large Data"
    
    # Header
    ws['A1'] = 'ID'
    ws['B1'] = 'Value'
    
    # 100 rows of data
    for i in range(2, 102):
        ws[f'A{i}'] = i - 1
        ws[f'B{i}'] = f'Item {i - 1}'
    
    wb.save(FIXTURE_DIR / 'large.xlsx')
    print(f"Created: {FIXTURE_DIR / 'large.xlsx'}")


if __name__ == '__main__':
    create_sample_xlsx()
    create_multi_sheet_xlsx()
    create_formulas_xlsx()
    create_large_xlsx()
    print("Done!")
