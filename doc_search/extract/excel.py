"""
Excel (.xlsx) text extraction using vendored openpyxl library.

Extracts text content from Excel workbooks with:
- One document per worksheet
- Header row detection for contextual extraction
- Cell values extracted row-by-row
- Metadata including row/column counts
"""

import sys
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from urllib.parse import quote

from ..core.paths import VENDOR_DIR
if str(VENDOR_DIR) not in sys.path:
    sys.path.insert(0, str(VENDOR_DIR))

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ExcelExtractor:
    """
    Extract text from Excel .xlsx files using openpyxl.
    
    Creates one document per worksheet with:
    - Header row detection for context-aware text extraction
    - Cell values extracted row-by-row
    - Metadata including row/column counts
    """
    
    def __init__(self, first_row_is_header: bool = True, max_rows: Optional[int] = None):
        """
        Initialize Excel extractor.
        
        Args:
            first_row_is_header: If True, treat first row as column headers
            max_rows: Maximum rows to extract per sheet (None = no limit)
        """
        self.first_row_is_header = first_row_is_header
        self.max_rows = max_rows
    
    def _format_cell_value(self, value: Any) -> str:
        """Convert cell value to string."""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'TRUE' if value else 'FALSE'
        if isinstance(value, (int, float)):
            # Format numbers nicely (avoid scientific notation for reasonable numbers)
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)
        import re
        text = str(value).strip()
        # Remove control characters that break URL handling and display
        text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
        return text
    
    def _extract_sheet(
        self,
        sheet,
        sheet_name: str,
        sheet_index: int,
        file_path: Path,
        workbook_name: str
    ) -> Dict[str, Any]:
        """
        Extract data from a single worksheet.
        
        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet
            sheet_index: Index of the sheet (0-based)
            file_path: Path to the workbook file
            workbook_name: Name of the workbook file
            
        Returns:
            Document dict with url, title, text, headings, metadata, error
        """
        result = {
            'url': f"file://{file_path.absolute()}#{quote(sheet_name)}",
            'title': f"{workbook_name} - {sheet_name}",
            'text': '',
            'headings': [],
            'metadata': {
                'doc_type': 'xlsx',
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'workbook': workbook_name,
                'row_count': 0,
                'col_count': 0,
                'file_path': str(file_path.absolute()),
            },
            'error': None
        }
        
        try:
            # Get dimensions
            if sheet.max_row is None or sheet.max_column is None:
                # Empty sheet
                return result
            
            row_count = sheet.max_row
            col_count = sheet.max_column
            
            # Apply max_rows limit if set
            if self.max_rows and row_count > self.max_rows:
                row_count = self.max_rows
            
            result['metadata']['row_count'] = row_count
            result['metadata']['col_count'] = col_count
            
            # Build merged cell map: (row, col) → value from top-left cell
            # Merged cells return None for all but the primary cell in openpyxl
            # Note: merged_cells is not available in read_only mode
            merged_map = {}
            try:
                for merge_range in sheet.merged_cells.ranges:
                    top_left = sheet.cell(merge_range.min_row, merge_range.min_col).value
                    for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
                        for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                            if row_idx == merge_range.min_row and col_idx == merge_range.min_col:
                                continue
                            merged_map[(row_idx, col_idx)] = top_left
            except AttributeError:
                # read_only worksheets don't support merged_cells
                pass
            
            # Extract all rows, filling in merged cell values
            rows = []
            if merged_map:
                # Need cell objects for row/column info
                for row in sheet.iter_rows(min_row=1, max_row=row_count, values_only=False):
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(self._format_cell_value(cell.value))
                        elif (cell.row, cell.column) in merged_map:
                            row_values.append(self._format_cell_value(merged_map[(cell.row, cell.column)]))
                        else:
                            row_values.append('')
                    rows.append(row_values)
            else:
                # No merged cells — use values_only for better performance
                for row in sheet.iter_rows(min_row=1, max_row=row_count, values_only=True):
                    row_values = [self._format_cell_value(cell) for cell in row]
                    rows.append(row_values)
            
            if not rows:
                return result
            
            # Determine headers
            headers = None
            start_row = 0
            if self.first_row_is_header and rows:
                headers = rows[0]
                start_row = 1
                # Add headers as headings (H2)
                for h in headers:
                    if h:
                        result['headings'].append((2, h))
            
            # Format text content
            lines = []
            for row in rows[start_row:]:
                # Skip empty rows
                if not any(cell for cell in row):
                    continue
                
                if headers:
                    # Format with headers: "Header1: value1, Header2: value2"
                    parts = []
                    for i, cell in enumerate(row):
                        if cell:
                            header = headers[i] if i < len(headers) and headers[i] else f"Col{i+1}"
                            parts.append(f"{header}: {cell}")
                    if parts:
                        lines.append(', '.join(parts))
                else:
                    # Format without headers: "value1 | value2 | value3"
                    non_empty = [cell for cell in row if cell]
                    if non_empty:
                        lines.append(' | '.join(non_empty))
            
            result['text'] = '\n'.join(lines)
            
        except Exception as e:
            result['error'] = f"Sheet extraction error: {e}"
        
        return result
    
    def extract_from_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Extract documents from an Excel file (one per sheet).
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            List of document dicts, one per sheet
        """
        file_path = Path(file_path)
        documents = []
        
        try:
            # Load workbook (normal mode to support merged cells, data_only to get values not formulas)
            wb = load_workbook(
                filename=str(file_path),
                read_only=False,
                data_only=True  # Get computed values instead of formulas
            )
            
            workbook_name = file_path.name
            
            for idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                doc = self._extract_sheet(
                    sheet=sheet,
                    sheet_name=sheet_name,
                    sheet_index=idx,
                    file_path=file_path,
                    workbook_name=workbook_name
                )
                documents.append(doc)
            
            wb.close()
            
        except InvalidFileException as e:
            documents.append({
                'url': f"file://{file_path}",
                'title': file_path.name,
                'text': '',
                'headings': [],
                'error': f'Invalid xlsx file: {e}',
                'metadata': {
                    'doc_type': 'xlsx',
                    'file_path': str(file_path)
                }
            })
        except Exception as e:
            documents.append({
                'url': f"file://{file_path}",
                'title': file_path.name,
                'text': '',
                'headings': [],
                'error': f'Extraction error: {e}',
                'metadata': {
                    'doc_type': 'xlsx',
                    'file_path': str(file_path)
                }
            })
        
        return documents
    
    def extract(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Alias for extract_from_file for consistency with other extractors.
        """
        return self.extract_from_file(file_path)


def extract_excel_text(file_path: str, **kwargs) -> str:
    """
    Convenience function to extract just the text from an Excel file.
    
    Args:
        file_path: Path to Excel file
        **kwargs: Passed to ExcelExtractor
        
    Returns:
        Extracted text from all sheets concatenated, or empty string on failure
    """
    extractor = ExcelExtractor(**kwargs)
    documents = extractor.extract(Path(file_path))
    
    texts = []
    for doc in documents:
        if doc.get('text') and not doc.get('error'):
            texts.append(f"=== {doc['title']} ===\n{doc['text']}")
    
    return '\n\n'.join(texts)


# CLI for testing
if __name__ == '__main__':
    import argparse
    import json
    
    parser = argparse.ArgumentParser(description='Extract text from Excel files')
    parser.add_argument('file', help='Path to .xlsx file')
    parser.add_argument('--json', action='store_true', help='Output as JSON')
    parser.add_argument('--no-headers', action='store_true', 
                        help='Do not treat first row as headers')
    parser.add_argument('--max-rows', type=int, default=None,
                        help='Maximum rows to extract per sheet')
    args = parser.parse_args()
    
    extractor = ExcelExtractor(
        first_row_is_header=not args.no_headers,
        max_rows=args.max_rows
    )
    file_path = Path(args.file)
    
    if not file_path.exists():
        print(f"Error: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)
    
    documents = extractor.extract(file_path)
    
    if args.json:
        # Convert headings tuples to lists for JSON
        for doc in documents:
            doc['headings'] = [list(h) for h in doc['headings']]
        print(json.dumps(documents, indent=2))
    else:
        for doc in documents:
            if doc['error']:
                print(f"Error: {doc['error']}", file=sys.stderr)
                continue
            
            print(f"Title: {doc['title']}")
            print(f"URL: {doc['url']}")
            
            meta = doc['metadata']
            print(f"Sheet: {meta['sheet_name']} ({meta['row_count']} rows, {meta['col_count']} cols)")
            
            if doc['headings']:
                print(f"\nHeaders ({len(doc['headings'])}):")
                for level, text in doc['headings'][:10]:
                    print(f"  [H{level}] {text[:60]}")
            
            print(f"\n{'-'*50}\n")
            preview = doc['text'][:1500]
            print(preview)
            if len(doc['text']) > 1500:
                print(f"\n... ({len(doc['text']) - 1500} more characters)")
            print()
